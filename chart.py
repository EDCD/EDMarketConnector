#!/usr/bin/python
#
# Creates an Excel spreadsheet graphing player stats
#
# Requires openpyxl >= 2.3
#

try:
    import lxml._elementpath	# Explicit dependency for py2exe
    import openpyxl
    if map(int, openpyxl.__version__.split('.')[:2]) < [2,3]:
        raise ImportError()
    have_openpyxl = True
except:
    have_openpyxl = False

import datetime
import time
import re

from os.path import isfile, join
from config import config


dataseries = [
    {
        'name':    'Combat',
        'axes':    ['Quantity', 'Profit [CR]'],
        'keys':    [
            ('Bounties',              ['stats', 'combat', 'bounty', 'qty']),
            ('Profit from bounties',  ['stats', 'combat', 'bounty', 'value'], True),
            ('Bonds',                 ['stats', 'combat', 'bond', 'qty']),
            ('Profit from bonds',     ['stats', 'combat', 'bond', 'value'], True),
            ('Assassinations',        ['stats', 'missions', 'assassin', 'missionsCompleted']),
            ('Profit from assassin',  ['stats', 'missions', 'assassin', 'creditsEarned'], True),
            ('Hunting',               ['stats', 'missions', 'bountyHunter', 'missionsCompleted']),
            ('Profit from hunting',   ['stats', 'missions', 'bountyHunter', 'creditsEarned'], True),
        ],
    },
    {
        'name':    'Trade',
        'axes':    ['Quantity', 'Profit [CR]'],
        'keys':    [
            ('Profit from trading',   ['stats', 'trade', 'profit'], True),
            ('Commodities traded',    ['stats', 'trade', 'qty']),
            ('Profit from smuggling', ['stats', 'blackMarket', 'profit'], True),
            ('Commodities smuggled',  ['stats', 'blackMarket', 'qty']),
            ('Profit from mining',    ['stats', 'mining', 'profit'], True),
            ('Fragments mined',       ['stats', 'mining', 'qty']),
            ('Fragments converted',   ['stats', 'mining', 'converted', 'qty']),
        ],
    },
    {
        'name':    'Explorer',
        'axes':    ['Quantity', 'Profit [CR]'],
        'keys':    [
            ('Profits from exploration',      ['stats', 'explore', 'creditsEarned'], True),
            ('Discovery scans',               ['stats', 'explore', 'scanSoldLevels', 'lev_0']),
            ('Level 2 detailed scans',        ['stats', 'explore', 'scanSoldLevels', 'lev_1']),
            ('Level 3 detailed scans',        ['stats', 'explore', 'scanSoldLevels', 'lev_2']),
            ('Bodies first discovered',       ['stats', 'explore', 'bodiesFirstDiscovered']),
            ('Hyperspace jumps',              ['stats', 'explore', 'hyperspaceJumps']),
        ]
    },
    {
        'name':    'Crime',
        'axes':    ['Quantity', 'Profit [CR]'],
        'keys':    [
            ('Fines',                 ['stats', 'crime', 'fine', 'qty']),
            ('Lifetime fine value',   ['stats', 'crime', 'fine', 'value'], True),
            ('Bounties',              ['stats', 'crime', 'bounty', 'qty']),
            ('Lifetime bounty value', ['stats', 'crime', 'bounty', 'value'], True),
            ('Profit from cargo',     ['stats', 'crime', 'stolenCargo', 'value'], True),
            ('Stolen cargo',          ['stats', 'crime', 'stolenCargo', 'qty']),
            ('Profit from goods',     ['stats', 'stolenGoods', 'profit'], True),
            ('Stolen goods',          ['stats', 'stolenGoods', 'qty']),
        ],
    },
    {
        'name':    'NPC',
        'axes':    'Quantity',
        'prefix':  ['stats', 'NPC', 'kills', 'ranks'],
        'keys':    [('Harmless', 'r0'), ('Mostly Harmless', 'r1'), ('Novice', 'r2'), ('Competent', 'r3'), ('Expert', 'r4'), ('Master', 'r5'), ('Dangerous', 'r6'), ('Deadly', 'r7'), ('Elite', 'r8'), ('Capital', 'rArray')],
    },
    {
        'name':    'PVP',
        'axes':    'Quantity',
        'prefix':  ['stats', 'PVP', 'kills', 'ranks'],
        'keys':    [('Harmless', 'r0'), ('Mostly Harmless', 'r1'), ('Novice', 'r2'), ('Competent', 'r3'), ('Expert', 'r4'), ('Master', 'r5'), ('Dangerous', 'r6'), ('Deadly', 'r7'), ('Elite', 'r8'), ('Capital', 'rArray')],
    },
    {
        'name':    'Balance',
        'axes':    ['Quantity', '[CR]'],
        'keys':    [
            ('Current balance',       ['commander', 'credits'], True),
            ('Spent on ships',        ['stats', 'ship', 'spend', 'ships'], True),
            ('Spent on outfitting',   ['stats', 'ship', 'spend', 'modules'], True),
            ('Spent on repairs',      ['stats', 'ship', 'spend', 'repair'], True),
            ('Insurance claims',      ['stats', 'ship', 'insurance', 'claims']),
            ('Total claim costs',     ['stats', 'ship', 'insurance', 'value'], True),
        ],
    },
    # {
    #     'name':    'Vanish',
    #     'prefix':  ['stats', 'vanishCounters'],
    #     'keys':    ['amongPeers', 'inDanger', 'inDangerWithPeers', 'isNotDying', 'noPeers', 'notInDanger'],
    # },
]


def export(data, csv=False):

    if not have_openpyxl: return False

    TITLE_F = openpyxl.styles.Font(bold=True)
    TITLE_A = openpyxl.styles.Alignment(horizontal='right')

    querytime = config.getint('querytime') or int(time.time())

    filename = join(config.get('outdir'), 'Cmdr %s.xlsx' % re.sub(r'[\\/:*?"<>|]', '_', data['commander']['name']))
    if not isfile(filename):
        wb = openpyxl.Workbook()
        try:
            wb.active.title = 'Combat'	# Workbook is created with one sheet - rename it
        except:
            pass			# except that it isn't under 2.30b1
    else:
        wb = openpyxl.load_workbook(filename)

    for thing in dataseries:
        legends = [x[0] for x in thing['keys']]
        keys    = [x[1] for x in thing['keys']]
        if thing.get('axes') and isinstance(thing['axes'], (list, tuple)):
            y2_axis = [len(x)>2 and x[2] for x in thing['keys']]
        else:
            y2_axis = None

        if thing['name'] in wb:
            ws = wb[thing['name']]
        else:
            ws = wb.create_sheet(title=thing['name'])

        # Add header row
        if ws.max_row <= 1:	# Returns 1 for empty sheet
            ws.append(['Date'] + legends)
            for i in range(ws.max_column):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i+1)].width = 17
            for row in ws.get_squared_range(1, 1, ws.max_column, 1):
                for cell in row:
                    cell.font = TITLE_F
                    cell.alignment = TITLE_A

        # Add data row
        vals = [datetime.datetime.fromtimestamp(querytime)]
        mydata = data
        if thing.get('prefix'):
            for key in thing['prefix']:
                mydata = mydata[key]
        for key2 in keys:
            if isinstance(key2, basestring):
                vals.append(mydata.get(key2, 0))
            else:
                value = mydata
                for key in key2:
                    value = value.get(key, 0)
                    if not value: break
                vals.append(value)
        ws.append(vals)
        ws.cell(row=ws.max_row, column=1).number_format = 'yyyy-mm-dd hh:mm:ss'	# just a string, not a style

        dates = openpyxl.chart.Reference(ws, 1, 2, 1, ws.max_row)

        chart = openpyxl.chart.ScatterChart()
        chart.title = thing['name']
        chart.width, chart.height = 60, 30	# in cm!
        chart.scatterStyle = 'lineMarker'
        chart.set_categories(dates)
        chart.x_axis.number_format = ('yyyy-mm-dd')	# date only
        chart.x_axis.majorGridlines = None

        if y2_axis:
            chart.y_axis.majorGridlines = None		# prefer grid lines on secondary axis
            chart2 = openpyxl.chart.ScatterChart()
            chart2.scatterStyle = 'lineMarker'
            # Hack - second chart must have different axis ID
            chart2.y_axis = openpyxl.chart.axis.NumericAxis(axId=30, crossAx=10, axPos='r', crosses='max')

        for i in range(len(keys)):
            series = openpyxl.chart.Series(openpyxl.chart.Reference(ws, i+2, 1, i+2, ws.max_row), dates, title_from_data=True)
            series.marker.symbol = 'diamond'
            if y2_axis and y2_axis[i]:
                series.dLbls = openpyxl.chart.label.DataLabels([openpyxl.chart.label.DataLabel(idx=ws.max_row-2, dLblPos='r', showSerName=True)])
                chart2.series.append(series)
            else:
                series.dLbls = openpyxl.chart.label.DataLabels([openpyxl.chart.label.DataLabel(idx=0, dLblPos='l', showSerName=True)])
                chart.series.append(series)
        if y2_axis:
            chart.y_axis.title  = thing['axes'][0]
            chart2.y_axis.title = thing['axes'][1]
            chart.z_axis = chart2.y_axis
            chart += chart2
        elif thing['axes']:
            chart.y_axis.title = thing['axes']
        ws.add_chart(chart, 'B2')

    wb.save(filename)
